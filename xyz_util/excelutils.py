#!/usr/bin/env python
# -*- coding:utf-8 -*-
import re

from six import text_type, string_types

try:
    from StringIO import StringIO
except:
    from io import StringIO

import logging

log = logging.getLogger("django")

import xlrd, xlwt


def coordinate_range_to_tuple(s):
    from openpyxl.utils import coordinate_to_tuple
    ps = s.split(':')
    p0 = coordinate_to_tuple(ps[0])
    p1 = coordinate_to_tuple(ps[1])
    return (p0[0] - 1, p1[0], p0[1] - 1, p1[1])


def get_grid_dict(excel_file, sheet_name):
    """
    把excel数据转成字典，以x,y坐标为key, 单元格值为value, 合并单元格的值会填充到每个子单元格
    """
    data = xlrd.open_workbook(excel_file, formatting_info=True)
    point_dict = {}
    table = data.sheet_by_name(sheet_name)
    for i in range(table.nrows):
        row = table.row_values(i)
        for j in range(table.ncols):
            if row[j] != "":
                point_dict[(i, j)] = row[j]
    for item in table.merged_cells:
        baser = item[0]
        basec = item[2]
        value = point_dict.get((baser, basec))
        if value is None:
            continue
        for i in range(item[0], item[1]):
            for j in range(item[2], item[3]):
                point_dict[(i, j)] = value
    return point_dict


def get_grid_dict_xlsx(excel_file, sheet_name=None):
    """
    把excel数据转成字典，以x,y坐标为key, 单元格值为value, 合并单元格的值会填充到每个子单元格
    """
    import openpyxl
    wb = openpyxl.load_workbook(excel_file)
    point_dict = {}
    ws = wb.get_sheet_by_name(sheet_name) if sheet_name else wb.active
    rs = ws.rows
    for i in range(ws.max_row):
        row = rs.next()
        for j in range(ws.max_column):
            v = row[j].value
            if v != "" and v is not None:
                point_dict[(i, j)] = v
    for mcr in ws.merged_cell_ranges:
        item = coordinate_range_to_tuple(mcr)
        baser = item[0]
        basec = item[2]
        value = point_dict.get((baser, basec))
        if value is None:
            continue
        for i in range(item[0], item[1]):
            for j in range(item[2], item[3]):
                point_dict[(i, j)] = value
    return point_dict


def filter_sheets_by_name(excel_file, re_sheet_name):
    """
    通过表名正则表达式，查找匹配的excel数据表
    """
    data = xlrd.open_workbook(excel_file, formatting_info=True)
    sheet_names = data.sheet_names()
    names = []
    for sheet_name in sheet_names:
        if re_sheet_name.match(sheet_name):
            names.append(sheet_name)
    return names


class Reader(object):

    def __init__(self, excel, row_top=0, field_names_template=[], min_fields_count=1, col_name_formater=lambda c: c):
        if isinstance(excel, string_types):
            self.workbook = xlrd.open_workbook(excel)
        elif isinstance(excel, xlrd.Book):
            self.workbook = excel
        elif hasattr(excel, "read"):
            self.workbook = xlrd.open_workbook(file_contents=excel.read())
        else:
            raise ValueError("excel bad format.")

        self.field_names_template = field_names_template
        self.row_top = row_top
        self.min_fields_count = min_fields_count
        self.col_name_formater = col_name_formater

    def to_list(self):
        res = []
        for sheet in self.workbook.sheets():
            nrows = sheet.nrows
            ncols = sheet.ncols

            if self.field_names_template:
                max_fields_count = 0
                for i in range(min(nrows, 10)):
                    line = sheet.row_values(i)
                    count = len([col for col in line if self.col_name_formater(col) in self.field_names_template])
                    if count > max_fields_count:
                        max_fields_count = count
                        row_top = i
                if max_fields_count < self.min_fields_count:
                    continue
            field_names = [self.col_name_formater(col) for col in sheet.row_values(row_top)]

            for i in range(row_top + 1, nrows):
                line = sheet.row_values(i)
                obj = dict([(field_names[col], line[col]) for col in range(0, ncols) if field_names[col]])
                vc = len([v for v in obj.values() if v])
                if vc >= self.min_fields_count:
                    res.append(obj)
        return res


def pandas_read(excel, trim_null_rate=0.5):
    from .pandasutils import trim, split_blocks, dataframe_to_table
    import pandas as pd
    try:
        sheets = pd.read_excel(excel, sheetname=None, header=None)
    except TypeError as e:
        sheets = pd.read_excel(excel, sheet_name=None, header=None)
    rs = []
    for s, df in sheets.items():
        bs = []
        for bl in split_blocks(df):
            df = trim(bl, null_rate=trim_null_rate)
            bs.append(dataframe_to_table(df))
        rs.append({'name': s, 'blocks': bs})
    return {'sheets': rs}


def excel2json(excel, row_top=0, field_names_template=[], min_fields_count=1, col_name_formater=lambda c: text_type(c)):
    """
    把excel数据转成json字典，本函数在实现非固定字段导入时很有用。
    比如excel：
         姓名 年龄  学号  辅导员
         张三 19   201301 李湘
         李四 20   201302

    excel2json(excel,field_names_template=["姓名","年龄","学号"],min_fields_count=3)

    执行结果：
        [
         {"姓名":"张三","年龄":19,"学号":"201301","辅导员":"李湘"},
         {"姓名":"李四","年龄":20,"学号":"201302"}
        ]

    :param excel: excel文件, 支持多种参数形式： string(文件路径), xlrd.Book, file object(支持.read()的类File对象)
    :param row_top: 指定表头行号，如果不指定field_names_template进行自动识别的话
    :param field_names_template: 表头可能包括的各个字段名，每个sheet会尝试从前10行里自动找出包含最多字段的行当作表头
        数据从表头以下的行开始读取
    :param min_fields_count: 至少要包括min_fields_count个字段才能认为是一个合格的表头，如果当前sheet找不到合格的表头，
        则认为当前sheet没有正式数据，直接忽略
    :return: 字典数组， 多个sheet的数据会合并到一个数组里，字典以表头字段名为key
    """
    if isinstance(excel, string_types):
        workbook = xlrd.open_workbook(excel)
    elif isinstance(excel, xlrd.Book):
        workbook = excel
    elif hasattr(excel, "read"):
        workbook = xlrd.open_workbook(file_contents=excel.read())
    res = []
    if isinstance(field_names_template, string_types):
        field_names_template = field_names_template.split(',')

    # ftc = len(field_names_template)
    def count_target_column(col):
        return sum([(1 if fn in col else 0) for fn in field_names_template])

    for sheet in workbook.sheets():
        nrows = sheet.nrows
        if nrows <= 0:
            continue
        ncols = sheet.ncols

        if field_names_template:
            max_fields_count = 0
            for i in range(min(nrows, 10)):
                line = sheet.row_values(i)
                count = len([col for col in line if count_target_column(col_name_formater(col))])
                if count > max_fields_count:
                    max_fields_count = count
                    row_top = i
            if max_fields_count < min_fields_count:
                continue
        field_names = [col_name_formater(col) for col in sheet.row_values(row_top)]

        for i in range(row_top + 1, nrows):
            line = sheet.row_values(i)
            obj = dict([(field_names[col], line[col]) for col in range(0, ncols) if field_names[col]])
            vc = len([v for v in obj.values() if v])
            if vc >= min_fields_count:
                res.append(obj)
    return res


class TableReader(object):

    def __init__(self, field_words=[]):
        self.fields = [
            (a, [f.strip() for f in b.split(',') if f.strip()])
            for a, b in field_words
        ]

    def recognize(self, ts):
        d = {}
        for fn, ws in self.fields:
            rs = []
            for i, t in enumerate(ts):
                c = 0
                for w in ws:
                    if w in t:
                        c += 1
                rs.append((c, i))
            rs.sort()
            top = rs[-1]
            if top[0] >0:
                d[fn] = ts[top[1]]
        return d

    def transform(self, ds):
        fm = self.recognize(list(ds[0].keys()))
        rs = []
        for d in ds:
            nd = dict([(k, d.get(v, None)) for k, v in fm.items()])
            rs.append(nd)
        return rs

    def read(self, excel_file):
        fts = reduce(lambda a, b: a + b, map(lambda a: a[1], self.fields), [])
        ds = excel2json(excel_file, field_names_template=fts)
        return self.transform(ds)

try:
    from excel_response import ExcelResponse

    class ExcelDumpsMixin(object):
        """
            与ListView一起，实现列表数据的excel导出功能。
            request.GET.get("format_") == "xls"  时，就会触发导出动作
        """
        xls_fields = []
        xls_headers = []
        force_csv = False
        xls_filename = "data"
        background_upto = 100000

        def get_xls_headers(self):
            hs = self.xls_headers
            if not hs:
                hs = []
                md = self.object_list.model
                from .modelutils import get_related_field
                fs = self.get_xls_fields()
                for f in fs:
                    try:
                        hs.append(get_related_field(md, f).verbose_name)
                    except:
                        hs.append(f.split(".")[-1])
            return hs

        def get_xls_fields(self):
            return self.xls_fields

        def get_xls_filename(self):
            return self.xls_filename

        def export_to_xls_in_background(self):
            from django.http import HttpResponseRedirect
            from django.shortcuts import resolve_url
            qset = self.object_list
            from django.contrib.contenttypes.models import ContentType
            params = dict(content_type_id=ContentType.objects.get_for_model(qset.model).id,
                          object_ids=list(qset.values_list("id", flat=True)),
                          accessors=self.get_xls_fields(),
                          headers=self.get_xls_headers())
            from xyz_common.models import ExcelTask
            task = ExcelTask()
            task.content_object = self.master
            task.owner = self.master.user
            task.params = params
            task.name = self.title
            task.status = 0
            task.save()
            return HttpResponseRedirect(resolve_url("common_config:excel-task-detail", pk=task.id))

        def export_to_xls(self):
            r = [self.get_xls_headers()]
            oc = self.object_list.count()
            log.info("start to prepare excel data:%s", oc)
            from django_tables2.utils import A
            accessories = [A(f) for f in self.get_xls_fields()]
            from .modelutils import get_object_accessor_value
            for item in self.object_list:
                r.append([get_object_accessor_value(item, a) for a in accessories])
            log.info("start to dumps excel:%s", len(r))
            return ColorExcelResponse(r, str(self.get_xls_filename()), force_csv=self.force_csv)

        def render_to_response(self, context, **response_kwargs):
            if self.request.GET.get("format_") == "xls":
                if self.object_list.count() > self.background_upto:
                    return self.export_to_xls_in_background()
                return self.export_to_xls()
            return super(ExcelDumpsMixin, self).render_to_response(context, **response_kwargs)


    class TablesExcelDumpsMixin(ExcelDumpsMixin):
        def get_xls_fields(self):
            t = self.get_table()
            return [c.accessor for f, c in t.columns.items()]

        def get_xls_headers(self):
            t = self.get_table()
            return [c.verbose_name for f, c in t.columns.items()]


    class ColorExcelResponse(ExcelResponse):
        def write_row(self, row):
            for colx, value in enumerate(row):
                if value is None and self.blanks_for_none:
                    value = ''

                if self.is_user_defined_class(value):
                    value = str(value)

                cell_style = self.styles['default']
                if type(value).__name__ in self.styles:
                    cell_style = self.styles[type(value).__name__]

                elif isinstance(value, string_types):
                    leading_zero_number_regex = re.compile(
                        r'^-?[0]+[0-9,]*$'
                    )
                    comma_separated_number_regex = re.compile(
                        r'^-?[0-9,]*\.[0-9]*$'
                    )
                    dollar_regex = re.compile(
                        r'^\$[0-9,\.]+$'
                    )

                    try:
                        if leading_zero_number_regex.match(value):
                            cell_style = xlwt.easyxf(
                                num_format_str='0' * len(value))
                        elif comma_separated_number_regex.match(value) and value != '-':
                            value = float(value.replace(',', ''))
                            if len(str(value)) > 15:
                                value = str(value)
                                cell_style = xlwt.easyxf(
                                    num_format_str='0' * len(value))
                        elif dollar_regex.match(value) and value != '-':
                            value = float(re.sub(r'[,$]', '', value))
                            cell_style = self.styles['currency']
                    except ValueError:
                        pass

                style = self.get_extra_style(self.rowx, colx, cell_style)
                self.sheet.write(self.rowx, colx, value, style=style)
                if self.auto_adjust_width:
                    width = len(text_type(value)) * 256
                    if width > self.widths.get(colx, 0):
                        if width >= self.ROW_LIMIT:
                            width = self.ROW_LIMIT - 1
                        self.widths[colx] = width
                        self.sheet.col(colx).width = width
            self.rowx += 1

        def set_output(self, output):
            self.output = output
            self.book = xlwt.Workbook(encoding=self.encoding, style_compression=2)
            self.sheet = self.book.add_sheet(self.sheet_name)

            self.styles = {
                'datetime': xlwt.easyxf(num_format_str='yyyy-mm-dd hh:mm:ss'),
                'date': xlwt.easyxf(num_format_str='yyyy-mm-dd'),
                'time': xlwt.easyxf(num_format_str='hh:mm:ss'),
                'default': xlwt.easyxf("font: name SimSun"),
                'currency': xlwt.easyxf(num_format_str='[$$-409]#,##0.00;-[$$-409]#,##0.00')
            }
            self.widths = {}
            self.rowx = 0

        def save(self):
            self.book.save(self.output)

        def write_xls(self, data, headers=None):
            if headers:
                self.write_row(headers)
            for row in data:
                self.write_row(row)
            self.save()
            return self.output

        def get_extra_style(self, rowx, colx, style):
            if (rowx, colx) in self.style_map:
                nstyle = xlwt.XFStyle()
                nstyle.font = style.font
                nstyle.alignment = style.alignment
                pattern = xlwt.Pattern()
                pattern.pattern = xlwt.Pattern.SOLID_PATTERN
                pattern.pattern_fore_colour = 2
                nstyle.pattern = pattern
                return nstyle
            return style

        def ensure_encoding(self, s):
            if self.encoding != "utf8" and isinstance(s, string_types):
                return s.decode("utf8").encode(self.encoding)
            return s

        def write_csv(self, data, headers=None):
            import csv
            writer = csv.writer(self.output)
            if headers:
                writer.writerow([self.ensure_encoding(item) for item in headers])
            for row in data:
                writer.writerow([self.ensure_encoding(item) for item in row])
            return self.output

        def __init__(self, *args, **kwargs):
            self.style_map = kwargs.pop("style_map", {})
            return super(ColorExcelResponse, self).__init__(*args, **kwargs)

        @property
        def as_xls(self):
            self.set_output(StringIO())
            return self.write_xls(self.data, self.headers)

        @property
        def as_csv(self):
            self.set_output(StringIO())
            return self.write_csv(self.data, self.headers)
except:
    import traceback
    log.debug('import excel_response error: %s', traceback.format_exc())